import openpyxl


class ExcelUtil:
    file = "test_data\\testData.xlsx"

    @staticmethod
    def get_row_count(file, sheet_name):
        workbook = openpyxl.load_workbook(filename=file)
        sheet = workbook[sheet_name]
        return sheet.max_row

    @staticmethod
    def get_column_count(file, sheet_name):
        workbook = openpyxl.load_workbook(filename=file)
        sheet = workbook[sheet_name]
        return sheet.max_column

    @staticmethod
    def get_cell_data(file, sheet_name, row_num, col_num):
        workbook = openpyxl.load_workbook(filename=file)
        sheet = workbook[sheet_name]
        var = sheet.cell(row=row_num, column=col_num).value
        return var

    @staticmethod
    def set_cell_data(file, sheet_name, row_num, col_num, data):
        workbook = openpyxl.load_workbook(filename=file)
        sheet = workbook[sheet_name]
        sheet.cell(row=row_num, column=col_num).value = data
        workbook.save(file)
